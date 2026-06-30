# -*- coding: utf-8 -*-
"""
Created on Mon Aug 14 13:48:50 2023

@author: liansiqi,wangdi
"""

import numpy as np
import pandas as pd
import datetime
import time
from datetime import timedelta
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, colors, Alignment
import openpyxl

#%% 时间设定
countdown = -1
filedate = (datetime.date.today() + datetime.timedelta(days = countdown + 1)).strftime('%Y%m%d')
countdate = (datetime.date.today() + datetime.timedelta(days = countdown)).strftime('%Y%m%d')
countdate2 = (datetime.date.today() + datetime.timedelta(days=countdown))


def getDatetimeToday():
    t = datetime.date.today()  # date类型
    dt = datetime.datetime.strptime(str(t), '%Y-%m-%d')  # date转str再转datetime
    return dt


def getDatetimeYesterday():
    today = getDatetimeToday()  # datetime类型当前日期
    yesterday = today + timedelta(days=countdown)  # 减去倒数日
    return yesterday


countdate3 = str(getDatetimeYesterday())

#%% 数据源读取
# 1、源-新内容运营日报
file_content_operation = "新内容运营日报V9_" + filedate + '.xlsx'
contents_detailed = pd.read_excel(file_content_operation,sheet_name = '明细')
contents_detailed = contents_detailed.fillna(0)
contents_summary = pd.read_excel(file_content_operation,sheet_name = '汇总')
content_kpi_nocheat = pd.read_excel(file_content_operation,sheet_name = '轮播池KPI去作弊')


# 2、源-分工表
file_division_of_labor = "分工-页面.xlsx"
division_of_labor_ori = pd.read_excel(file_division_of_labor, sheet_name='Sheet1', index_col='日期')


# 3、源-留存率表
file_retention = "103104点击留存.xlsx"
click_retention_app = pd.read_excel(file_retention,sheet_name = "1-all-103104点击次留率")
click_retention_103104 = pd.read_excel(file_retention,sheet_name = "2-103104-103104点击次留")
click_retention_104 = pd.read_excel(file_retention,sheet_name = "3-104点击用户次留-104点击用户次留")
click_retention_104_pos = pd.read_excel(file_retention,sheet_name = "4-104各位置点击次留-104各位置点击次留")


# 4、辅助表：各池子表头
temp_table = pd.read_excel("各池子表头.xlsx")

#%% 常量序列
dimensions_order = ['日期','业务池','内容ID','标题','内容类型','调性','一级分类','二级分类','是否大事件卡片','是否屏蔽北京','入池时间','入池人','移除时间','移除人','发布时间']
dimensions_order_12 = ['日期','业务池','内容ID','标题','内容类型','调性','一级分类','二级分类','是否大事件卡片','是否屏蔽北京','入池时间','入池人','移除时间','移除人','发布时间','曝光位置-业务位置']

metrics = ['曝光量','曝光用户量', '曝光点击量','点击用户量','浏览次数', '浏览用户量','总浏览时长(s)','相关视频浏览时长(s)','汇总浏览时长','总次均浏览时长(s)','总次均浏览进度','总次均总浏览时长(s)\n-(汇总浏览时长/浏览次数)']

pos_detailed = ['置顶下1条-时间线','置顶下1条-大事件卡片','置顶下1条-非大事件卡片','置顶下2条-新闻资讯','置顶下2条-本地编辑内容','置顶下2条-视频']


table_order = ['日期', '业务池', '内容ID', '标题', '曝光量', '曝光用户量', '曝光点击量', '点击用户量', 'CTR(%)', '浏览次数', '浏览用户量', '总浏览时长(s)', '相关视频浏览时长(s)', '汇总浏览时长',
            '次均浏览时长(s)', '次均浏览进度', '次均总浏览时长(s)\n-(汇总浏览时长/浏览次数)', '内容类型', '调性', '一级分类', '二级分类', '是否大事件卡片', '是否屏蔽北京', '入池时间', '入池人', '移除时间', '移除人', '发布时间']

#%% 函数
# 筛选出计算日的所有数据
def get_countday_data(df,countdate,column_name):
    df = df[df[column_name] == int(countdate)]
    df = df.reset_index(drop=True)
    return df


# 率指标、次均指标统一清洗
def calculate(df):
    df['CTR(%)'] = round(df['曝光点击量'] / df['曝光量'], 4).apply(lambda x: format(x, '.2%'))
    df['次均浏览时长(s)'] = df['总次均浏览时长(s)'] / df['浏览次数']
    df['次均浏览进度'] = df['总次均浏览进度'] / df['浏览次数'] 
    df['次均总浏览时长(s)\n-(汇总浏览时长/浏览次数)'] = df['总次均总浏览时长(s)\n-(汇总浏览时长/浏览次数)'] / df['浏览次数']
    #df = df[table_order]
    df = df.sort_values(by=['曝光量'], ascending=False)
    return df

# 自动调整列宽
def adjust_column_dimension(ws, min_row, min_col, max_col):
    column_widths = []
    for i, col in enumerate(ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)):
        for cell in col:
            value = cell.value
            if value is not None:
                if isinstance(value, str) is False:
                    value = str(value)
                try:
                    column_widths[i] = max(column_widths[i], len(value))
                except IndexError:
                    column_widths.append(len(value))
    for i, width in enumerate(column_widths):
        col_name = get_column_letter(min_col + i)
        value = column_widths[i] + 6
        ws.column_dimensions[col_name].width = value


#%% 数据清洗
# 1、日期筛选
#contents_detailed = get_countday_data(contents_detailed, countdate,"日期")
contents_summary = get_countday_data(contents_summary, countdate,"日期")
content_kpi_nocheat = get_countday_data(content_kpi_nocheat, countdate,"日期")

# 2、轮播池明细数据
contents_detailed = contents_detailed.query("业务池 == '轮播池'").reset_index(drop=True)
contents_detailed['总次均浏览时长(s)'] = contents_detailed['次均浏览时长(s)'] * contents_detailed['浏览次数']
contents_detailed['总次均浏览进度'] = contents_detailed['次均浏览进度'] * contents_detailed['浏览次数']
contents_detailed['总次均总浏览时长(s)\n-(汇总浏览时长/浏览次数)'] = contents_detailed['次均总浏览时长(s)\n-(汇总浏览时长/浏览次数)'] * contents_detailed['浏览次数']

# 所有位置按dimensions_order进行聚合
contents_all = contents_detailed.groupby(dimensions_order)[metrics].sum().reset_index(drop=False)
#contents_all.to_excel("s1.xlsx")
contents_all = calculate(contents_all)
#contents_all.to_excel("s2.xlsx")
# 12位置按dimensions_order_12进行聚合
contents_12 = contents_detailed[contents_detailed['曝光位置-业务位置'].isin(pos_detailed)]
contents_12 = contents_12.groupby(dimensions_order_12)[metrics].sum().reset_index(drop=False)
contents_12 = calculate(contents_12)
# 其他位置按dimensions_order进行聚合
contents_others = contents_detailed[contents_detailed['曝光位置-业务位置'].isin(['其他'])]
contents_others = contents_others.groupby(dimensions_order)[metrics].sum().reset_index(drop=False)
contents_others = calculate(contents_others)


#%% 各池子数据
# 表1：整体数据
# Part1:具体各位置数据
thatday_summary_all = contents_summary[contents_summary['业务池'] == '置顶下轮播池'].reset_index(drop=True)
thatday_summary_12 = thatday_summary_all[thatday_summary_all['曝光位置-业务位置'].isin(['置顶下1条','置顶下2条','置顶下2条-新闻资讯','置顶下2条-本地编辑内容','置顶下2条-视频'])]
#thatday_summary_12 = thatday_summary_all[thatday_summary_all['曝光位置-业务位置'].isin(['置顶下1条','置顶下2条'])]
temp_table[['日期', '业务池', '曝光量', '曝光位置-业务位置', '曝光点击量', 'CTR(%)', '汇总的浏览时长', '次均浏览时长(s)', '次均浏览进度']] = thatday_summary_12[[
    '日期', '业务池', '曝光量', '曝光位置-业务位置', '曝光点击量', 'CTR(%)', '汇总的浏览时长', '次均浏览时长(s)', '次均浏览进度']]
temp_table = temp_table.sort_values(by='曝光位置-业务位置', ascending=True).reset_index(drop=True)

# Part2:置顶下12汇总数据
content_kpi_nocheat = content_kpi_nocheat.reset_index(drop=True)
temp_table['头条访问用户'][0] = round(int(content_kpi_nocheat.loc[0, '头条访问用户']), 0)
temp_table['头条访问用户'][1:5] = ""

# Part3:客户端核心数据
temp_table['客户端DAU'][0] = int(thatday_summary_all['去作弊DAU'][0])
temp_table['置顶下1或2人均浏览时长(s)'] = round(
    temp_table['汇总的浏览时长'] / temp_table['客户端DAU'][0], 2)
temp_table['置顶下1&2合计时长'][0] = temp_table['汇总的浏览时长'][0]+temp_table['汇总的浏览时长'][1]
temp_table['置顶下1&2人均浏览时长(s)'][0] = round(
    temp_table['置顶下1&2合计时长'][0]/temp_table['客户端DAU'][0], 2)
temp_table['每DAU阅读时长'][0] = round(float(content_kpi_nocheat.loc[0, '每DAU阅读时长']), 2)
temp_table['CTR(%)'] = round(temp_table['曝光点击量'] / temp_table['曝光量'],
                           4).apply(lambda x: format(x, '.2%'))

temp_table['客户端人均总时长(minute)'][0] = round(thatday_summary_all.iloc[0, 17], 2)
temp_table['每DAU人均使用时长（min)-去作弊'][0] = round(thatday_summary_all['每DAU人均使用时长（min)-去作弊'].iloc[0], 2)

thatday_summary_others = thatday_summary_all[thatday_summary_all['曝光位置-业务位置'].isin([
    '头条其它位置'])].reset_index(drop=True)

temp_table['103/104所有位置占客户端时长比例'][0] = format((
    (thatday_summary_others['汇总的浏览时长'][0]/temp_table['客户端DAU'][0]) + temp_table['置顶下1&2人均浏览时长(s)'][0])/(contents_summary.iloc[0, 17]*60), '.2%')
temp_table['置顶下1或2人均浏览时长(s)'][2:5] = ""
temp_table = temp_table.drop(['置顶下1&2人均浏览时长(s)', '置顶下1&2占客户端时长比例'], axis=1)


# Part4:103104点击留存
# 使用melt函数合并列
melted_click_retention_103104 = pd.melt(click_retention_103104, id_vars=['day'], var_name='位置', value_name='点击次留率')
# 定义属性值的映射关系
attr_map = {'103点击次留率': '置顶下1条', '104点击次留率': '置顶下2条'}
# 使用map函数修改属性值
melted_click_retention_103104['位置'] = melted_click_retention_103104['位置'].map(attr_map)

# 定义属性值的映射关系
#attr_map1 = {'104_news':'置顶下2条-新闻资讯','104_local:'置顶下2条-本地编辑内容','104_news':'置顶下2条-新闻资讯'}
# 使用map函数修改属性值
click_retention_104_pos['pos'] = click_retention_104_pos['pos'].replace({'104_news': '置顶下2条-新闻资讯', '104_local': '置顶下2条-本地编辑内容', '104_video': '置顶下2条-视频'})
click_retention_104_pos = click_retention_104_pos.rename(columns={'pos': '位置'})
click_retention_104_pos = click_retention_104_pos.sort_values(by='位置', ascending=True)
click_retention_combine = pd.concat([melted_click_retention_103104,click_retention_104_pos],axis=0)

merged_retention = pd.merge(click_retention_app,click_retention_combine, on='day', how='left')
merged_retention = merged_retention[['day','103104点击次留率','点击次留率','位置']]
merged_retention.loc[1:, '103104点击次留率'] = ''


# 表2：置顶下12位
thatday_summary_12 = thatday_summary_12.sort_values(by='曝光位置-业务位置', ascending=True).reset_index(drop=True)
thatday_summary_12 = thatday_summary_12[['日期', '曝光位置-业务位置', '头条访问UV', '点击用户量', '点击用户量/头条访问UV']]
thatday_summary_12['曝光位置-业务位置'][0] = '置顶下第1位'
thatday_summary_12['曝光位置-业务位置'][1] = '置顶下第2位'
thatday_summary_12['业务池'] = '置顶下轮播池'
thatday_summary_12['点击用户量/头条访问UV'] = thatday_summary_12['点击用户量/头条访问UV'].apply(lambda x: format(x, '.2%'))

#thatday_summary_12['点击用户量/头条访问UV'] = thatday_summary_12['点击用户量/头条访问UV'].apply(lambda x: format(x, '.2%'))
thatday_summary_12['置顶下1条2条：点击用户量/头条访问UV'] = thatday_summary_all[thatday_summary_all['曝光位置-业务位置'] == '置顶下1条2条']\
    .reset_index(drop=True)['点击用户量/头条访问UV'].apply(lambda x: format(x, '.2%'))[0]
thatday_summary_12['头条整体：点击用户量/头条访问UV'] = thatday_summary_all[thatday_summary_all['曝光位置-业务位置'] == '头条整体']\
    .reset_index(drop=True)['点击用户量/头条访问UV'].apply(lambda x: format(x, '.2%'))[0]
thatday_summary_12.loc[thatday_summary_12['曝光位置-业务位置'] == '置顶下2条', '置顶下1条2条：点击用户量/头条访问UV'] = np.nan
thatday_summary_12.loc[thatday_summary_12['曝光位置-业务位置'] == '置顶下2条', '头条整体：点击用户量/头条访问UV'] = np.nan

#thatday_summary_12.loc[[0, 1], '曝光位置-业务位置'] = ['置顶下第1位', '置顶下第2位']
thatday_summary_12 = thatday_summary_12[['日期','业务池','曝光位置-业务位置', '头条访问UV', '点击用户量', '点击用户量/头条访问UV','置顶下1条2条：点击用户量/头条访问UV','头条整体：点击用户量/头条访问UV']]
thatday_summary_12['置顶下1条2条：点击用户量/头条访问UV'][1] = ""
thatday_summary_12['头条整体：点击用户量/头条访问UV'][1] = ""
thatday_summary_12 = thatday_summary_12.iloc[:2, :]


# 表3：其他位置
thatday_summary_others = thatday_summary_others[['日期', '业务池',
                               '曝光位置-业务位置', '曝光量', '曝光点击量', 'CTR(%)', '汇总的浏览时长']]
thatday_summary_others['CTR(%)'] = round(thatday_summary_others['曝光点击量'] /
                              thatday_summary_others['曝光量'], 4).apply(lambda x: format(x, '.2%'))


#%% Sheet:编辑每日贡献时长
# Part1：编辑每日贡献时长
# 编辑分工
division_of_labor = (division_of_labor_ori.unstack().apply(lambda x: pd.Series(str(x).split(' '))).stack().reset_index().drop('level_2', axis=1))
division_of_labor.columns = ['date','班次','入池人']
division_of_labor = division_of_labor.dropna()
#division_of_labor = division_of_labor[division_of_labor['date'] == countdate3]
division_of_labor = division_of_labor[division_of_labor['date'] == countdate3].reset_index(drop=True)
editor_military = {'date':division_of_labor.iloc[[0],[0]].values[0][0],
            '班次':"104军事视频",
            '入池人':"姚文广"
    }
division_of_labor.loc[len(division_of_labor)] = editor_military


# 维度：编辑；指标：总浏览时长、贡献度、平均阅读进度
editor_contrib = contents_all[['日期', '入池人', '曝光量', '曝光点击量', '次均浏览进度', '汇总浏览时长']]
editor_contrib['总次均浏览进度'] = editor_contrib['曝光点击量'] * editor_contrib['次均浏览进度']
editor_contrib_sum = editor_contrib.groupby(['入池人'],as_index=False)[['曝光量','曝光点击量','总次均浏览进度','汇总浏览时长']].sum()
editor_contrib_sum['平均阅读进度'] = (editor_contrib_sum['总次均浏览进度'] / editor_contrib_sum['曝光点击量'] / 100).apply(lambda x: format(x, '.2%'))
editor_contrib_sum['贡献度'] = round(editor_contrib_sum['汇总浏览时长']**2 / editor_contrib_sum['曝光量'], 0)
editor_contrib_sum = editor_contrib_sum.sort_values(by='汇总浏览时长', ascending=False)
editor_contrib_sum = editor_contrib_sum[['入池人', '汇总浏览时长', '贡献度', '平均阅读进度']]
editor_contrib_sum.insert(loc=0, column='date', value=countdate3)
editor_contrib_sum['date'] = pd.to_datetime(editor_contrib_sum['date'],errors='coerce')

# 维度：编辑；指标：当日发布的内容数
editor_contentnum = contents_all[['入池人','入池时间','内容ID']]
print("-----------------------------------------------")
print(type(editor_contentnum['入池时间'].iloc[0]))
print("-----------------------------------------------")
editor_contentnum['入池时间'] = pd.to_datetime(editor_contentnum['入池时间'],errors='coerce').dt.date
editor_contentnum = editor_contentnum[(editor_contentnum['入池时间'] == countdate2)]
editor_contentnum = pd.DataFrame(editor_contentnum.groupby('入池人')['内容ID'].nunique())
editor_contentnum = editor_contentnum.rename(columns={'内容ID': '当天发布条数(去重)'})

editor_contrib_sum = editor_contrib_sum.merge(editor_contentnum, on="入池人", how="left")

# 汇总编辑每日贡献时长
editor_contrib_all = pd.merge(editor_contrib_sum, division_of_labor, on=['date', '入池人'], how='left')
editor_contrib_all['班次1'] = pd.Categorical(editor_contrib_all['班次'], [\
                                       '早班103位置', '下午班103位置', '早班104位置', '下午班104位置', '早班内容审核', '下午班内容审核','早班104视频辅助','下午班104视频辅助'])
editor_contrib_all.sort_values(by=['班次1', '汇总浏览时长'], ascending=[True, False], inplace=True)
editor_contrib_all = editor_contrib_all[['date','班次','入池人','汇总浏览时长','贡献度','平均阅读进度','当天发布条数(去重)']]


for i in range(editor_contrib_all.shape[0]):
    if editor_contrib_all.loc[i, '班次'] == np.nan:
        editor_contrib_all.loc[i, '当天发布条数(去重)'] = 0
        print(editor_contrib_all.loc[i, '入池人'])
    else:
        editor_contrib_all.loc[i,'当天发布条数(去重)'] = editor_contrib_all.loc[i, '当天发布条数(去重)']

editor_contrib_all = editor_contrib_all.replace(np.nan, '')
editor_contrib_all.drop(editor_contrib_all[editor_contrib_all['入池人'].isin([0])].index, inplace=True)


# Part2：当日各权重发布条数,但不显示权重60的数据
weight_contentnum = contents_detailed[['入池时间','权重','内容ID']]
weight_contentnum['入池时间'] = pd.to_datetime(weight_contentnum['入池时间'],errors='coerce').dt.date
weight_contentnum = weight_contentnum[(weight_contentnum['入池时间'] == countdate2)]
weight_contentnum = pd.DataFrame(weight_contentnum.groupby('权重')['内容ID'].nunique())
weight_contentnum = weight_contentnum.rename(columns={'内容ID': '发布条数'})
# 删除权重为60
weight_contentnum = weight_contentnum[(weight_contentnum.index) > 60]

# Part3：103104调性数据，维度：调性；指标：总浏览时长、曝光量
tonality_contentnum = contents_12[contents_12['调性'].isin(['左','中','右'])]
#tonality_contentnum.to_excel("调性明细2.xlsx")
tonality_contentnum = tonality_contentnum.groupby(['调性']).agg({'汇总浏览时长': sum, '曝光量': sum}).reset_index(drop=False)

# Part4：104视频数据，维度：班次，入池人
#video_detail_104 = contents_detailed[contents_detailed['曝光位置-业务位置'].isin(['置顶下2条-新闻资讯','置顶下2条-视频'])]
video_detail_104 = contents_12[contents_12['曝光位置-业务位置'].isin(['置顶下2条-新闻资讯','置顶下2条-视频'])]

video_detail_104 = video_detail_104.query("内容类型 == '视频'")
video_detail_104 ['入池日期'] = pd.to_datetime(video_detail_104 ['入池时间'],errors='coerce').dt.date
video_detail_104 = video_detail_104[(video_detail_104['入池日期'] == countdate2)]
video_detail_104_temp = video_detail_104[['内容ID', '入池人','曝光量', '曝光点击量','浏览次数', '次均浏览进度', '汇总浏览时长']]
video_detail_104_temp['总浏览进度'] = video_detail_104_temp['浏览次数'] * video_detail_104_temp['次均浏览进度']
video_editor = video_detail_104_temp.groupby(['入池人'])[['曝光量', '曝光点击量','浏览次数', '总浏览进度', '汇总浏览时长']].sum().reset_index(drop=False)
video_editor['CTR'] = video_editor['曝光点击量']/video_editor['曝光量']
video_editor['次均浏览进度'] = video_editor['总浏览进度']/video_editor['浏览次数']/100
video_editor['总浏览时长'] = video_editor['汇总浏览时长']
video_num = pd.DataFrame(video_detail_104.groupby('入池人')['内容ID'].nunique())
video_num = video_num.rename(columns={'内容ID': '当天发布条数(去重)'})
# 低质内容
video_detail_104['CTR'] = video_detail_104['曝光点击量']/video_detail_104['曝光量']
video_detail_104_low= video_detail_104.query("曝光量>1000 and CTR<0.09")
video_detail_104_low_result = video_detail_104_low.groupby(['入池人']).agg({'曝光量':sum,'曝光点击量':sum}).reset_index(drop = False)
video_detail_104_low_result['低质内容CTR'] = video_detail_104_low_result['曝光点击量']/video_detail_104_low_result['曝光量']
video_detail_104_low_result = video_detail_104_low_result.rename(columns={'曝光量':'低质曝光量','曝光点击量':'低质曝光点击量'})
video_paiban = editor_contrib_all[['班次','入池人']]
video_editor_result = video_editor.merge(video_paiban,on='入池人')
video_editor_result = video_editor_result.merge(video_num,on='入池人')
video_editor_result = video_editor_result.query("班次=='早班104位置' or 班次=='下午班104位置' or 入池人 == '杨竞' or 入池人 == '姚文广' or 班次=='早班104视频辅助' or 班次=='下午班104视频辅助'")
#print("--------:"+video_editor_result.columns)
video_editor_result = video_editor_result.merge(video_detail_104_low_result,on='入池人',how='left')
video_editor_result['低质曝光占比'] = video_editor_result['低质曝光量']/video_editor_result['曝光量']
video_editor_result = video_editor_result[['班次','入池人','CTR','总浏览时长','次均浏览进度','当天发布条数(去重)','低质曝光占比','低质曝光量','低质内容CTR']].sort_values(by=['班次'],ascending=False)
video_editor_result = video_editor_result.replace(np.nan,0)

#%% 104视频明细
def Video_data_processing(df):    
    release_time = df[['内容ID','发布时间']]
    release_time = release_time.sort_values(['内容ID','发布时间'],ascending=[True,True])
    release_time = release_time.drop_duplicates(subset='内容ID',keep ='first')
    tmp_df  = df.groupby(['日期','业务池','内容ID','标题','内容类型','入池时间','入池人','移除时间','移除人']).sum()[['曝光量','曝光点击量']]
    tmp_df = tmp_df.reset_index()    
    tmp_df['CTR(%)'] = round(tmp_df['曝光点击量'] / tmp_df['曝光量'], 4).apply(lambda x: format(x, '.2%'))
    tmp_df = pd.merge(left = tmp_df,right = release_time,on='内容ID')
    tmp_df = tmp_df[list(df.columns)]
    return tmp_df

video_detail_104 = video_detail_104[['日期','业务池','内容ID','标题','曝光量', '曝光点击量', 'CTR(%)', '内容类型', '入池时间', '入池人', '移除时间', '移除人', '发布时间']]
filename = str(countdate)+"-104视频明细.xlsx"
video_detail_104 = Video_data_processing(video_detail_104)
video_detail_104.to_excel(filename,index=0)

#%% 104点击留存
# 用户数统计
def click_video_num(data):
    df1 = data.loc[:, ['日期', '用户活跃度', '用户类型', '点击用户数']]
    df2 = df1.groupby(['日期', '用户类型']).sum('点击用户数').reset_index()
    df2['用户活跃度'] = '总计'
    df1 = df1.loc[data['用户活跃度'].isin(['高活', '中活', '7日内新用户']), :]
    #result = df1.append(df2)
    result = pd.concat([df1,df2],axis=0)
    result['用户类型'] = result['用户类型'].map({'点击过104视频': '104视频点击用户数',
                                        '未点击过104视频': '未点击104视频，但点了104其他内容的用户数'})
    result['用户类型'] = pd.Categorical(result['用户类型'].tolist(), categories=list([
                                    '104视频点击用户数', '未点击104视频，但点了104其他内容的用户数']), ordered=True)
    result['用户活跃度'] = pd.Categorical(result['用户活跃度'].tolist(), categories=list([
                                     '高活', '中活', '7日内新用户', '总计']), ordered=True)
   # result['用户类型'] = result['用户类型'].astype("category")
   # result['用户类型'].cat.set_categories(['104视频点击用户数','未点击104视频用户数'],inplace=True)
    #result['用户活跃度'] = result['用户活跃度'].astype("category")
    # result['用户活跃度'].cat.set_categories(['高活','中活','7日内新用户','总计'],inplace=True)

    result = result.sort_values(['日期', '用户类型', '用户活跃度']).reset_index().loc[:, [
        '日期', '用户类型', '用户活跃度', '点击用户数']]
    result = result.pivot_table(
        index='日期', columns=['用户类型', '用户活跃度'], values='点击用户数').fillna(0)
    return result

# 留存率统计
def click_video_retention(data):
    df1 = data.loc[data['用户活跃度'].isin(['高活', '中活', '7日内新用户']), :]
    df1 = df1.loc[:, ['日期', '用户活跃度', '用户类型', '点击用户数', '留存率']]
    df2 = df1.loc[:, ['日期', '用户活跃度', '用户类型', '留存率']]
    result = df2
    result['用户类型'] = result['用户类型'].map({'点击过104视频': '104视频点击用户客户端次留',
                                        '未点击过104视频': '未点击104视频，但点了104其他内容的用户客户端次留'})
    result['用户类型'] = pd.Categorical(result['用户类型'].tolist(), categories=list(
        ['104视频点击用户客户端次留', '未点击104视频，但点了104其他内容的用户客户端次留']), ordered=True)
    result['用户活跃度'] = pd.Categorical(result['用户活跃度'].tolist(
    ), categories=list(['高活', '中活', '7日内新用户']), ordered=True)
    result = result.sort_values(['日期', '用户类型', '用户活跃度']).reset_index().loc[:, [
        '日期', '用户类型', '用户活跃度', '留存率']]

    #result['留存率'] = result['留存率'].apply(lambda x: format(x, '.2%'))
    result = result.pivot_table(
        index='日期', columns=['用户类型', '用户活跃度'], values='留存率').fillna(0)
    for i in range(result.shape[1]):
        result.iloc[:, i] = result.iloc[:, i].apply(lambda x: format(x, '.2%'))
    return result

click_video_num = click_video_num(click_retention_104)
click_video_retention = click_video_retention(click_retention_104)
retention_104 = pd.merge(click_video_num, click_video_retention, on="日期")
retention_104 = retention_104.sort_index(ascending=False)
retention_104.index = retention_104.index.set_names('')
retention_104.columns = retention_104.columns.set_names({'用户活跃度': '日期'})


#%% 输出文件
writer = pd.ExcelWriter(str(countdate)+"内容运营表.xlsx", engine='xlsxwriter')
retention_104.to_excel(writer, sheet_name="104点击留存")
writer.sheets['104点击留存'].set_row(2, None, None, {'hidden': True})
writer.sheets['104点击留存']
# 内容明细
contents_all.to_excel(excel_writer=writer,sheet_name='所有位置' + str(countdate), index=0)
contents_all = writer.sheets['所有位置' + str(countdate)]
contents_all.set_column("D:D", 47)

contents_12.to_excel(excel_writer=writer,sheet_name='12位置' + str(countdate), index=0)
contents_12 = writer.sheets['12位置' + str(countdate)]
contents_12.set_column("D:D", 47)

contents_others.to_excel(excel_writer=writer,sheet_name='其他位置' + str(countdate), index=0)
contents_others = writer.sheets['其他位置' + str(countdate)]
contents_others.set_column("D:D", 47)
# 各池子数据
temp_table.to_excel(excel_writer=writer, sheet_name='各池子数据' + str(countdate), index=0)
thatday_summary_12.to_excel(excel_writer=writer, sheet_name='各池子数据' + str(countdate), startrow=temp_table.shape[0]+2, index=0)
thatday_summary_others.to_excel(excel_writer=writer, sheet_name='各池子数据' + str(countdate),
                     startrow=temp_table.shape[0] + thatday_summary_12.shape[0] + 4, index=0)
merged_retention.to_excel(excel_writer=writer, sheet_name='各池子数据' + str(countdate),
                     startrow=temp_table.shape[0] + thatday_summary_12.shape[0] + thatday_summary_others.shape[0] + 6, index=0)
# 编辑每日贡献时长
editor_contrib_all.to_excel(excel_writer=writer, sheet_name='编辑每日贡献时长' + str(countdate), index=0)
weight_contentnum.to_excel(excel_writer=writer, sheet_name='编辑每日贡献时长' + str(countdate), startrow=editor_contrib_all.shape[0]+3)
tonality_contentnum.to_excel(excel_writer=writer, sheet_name='编辑每日贡献时长' + str(
    countdate), startrow=editor_contrib_all.shape[0]+weight_contentnum.shape[0]+6,index=0)

video_editor_result.to_excel(excel_writer=writer, sheet_name='编辑每日贡献时长' + str(
    countdate), startrow=editor_contrib_all.shape[0]+weight_contentnum.shape[0]+tonality_contentnum.shape[0]+9,index=0)

#writer.save()
writer.close()


wb = openpyxl.load_workbook(str(countdate)+"内容运营表.xlsx")
ws = wb['104点击留存']
adjust_column_dimension(ws, 1, 1, ws.max_column)
wb.save(str(countdate)+"内容运营表.xlsx")
wb.close()


